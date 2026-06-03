import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Falta instalar openpyxl. Ejecutá: pip install openpyxl")
    raise SystemExit(1)

ARCHIVO_JSON = Path("productos.json")

POSIBLES_EXCEL = [
    "victoria_bath.xlsx",
    "victoria_secret_bath.xlsx",
    "productos_victoria_bath.xlsx",
    "productosEspeciales.xlsx",
    "productos_especiales.xlsx",
]

CATEGORIAS_VALIDAS = {
    "VICTORIA'S SECRET": "VICTORIA'S SECRET",
    "VICTORIA SECRET": "VICTORIA'S SECRET",
    "VICTORIAS SECRET": "VICTORIA'S SECRET",
    "VICTORIA´S SECRET": "VICTORIA'S SECRET",
    "VICTORIA’S SECRET": "VICTORIA'S SECRET",
    "BATH AND BODY WORKS": "BATH AND BODY WORKS",
    "BATH & BODY WORKS": "BATH AND BODY WORKS",
    "BATH BODY WORKS": "BATH AND BODY WORKS",
}


def limpiar_texto(valor):
    return str(valor or "").strip()


def normalizar_categoria(valor):
    categoria = limpiar_texto(valor).upper()
    return CATEGORIAS_VALIDAS.get(categoria)


def formatear_precio(valor):
    if valor is None or str(valor).strip() == "":
        return "Consultar precio"

    texto = str(valor).strip()
    if texto.lower() in ["consultar", "consultar precio", "sin precio"]:
        return "Consultar precio"

    texto_limpio = texto.replace("₡", "").replace("¢", "").replace(",", "").strip()

    try:
        numero = float(texto_limpio)
        return f"₡{numero:.2f}"
    except ValueError:
        return texto


def buscar_excel():
    if len(sys.argv) > 1:
        archivo = Path(sys.argv[1])
        if archivo.exists():
            return archivo
        print(f"No encontré el Excel indicado: {archivo}")
        raise SystemExit(1)

    for nombre in POSIBLES_EXCEL:
        archivo = Path(nombre)
        if archivo.exists():
            return archivo

    print("No encontré el Excel.")
    print("Usá uno de estos nombres:")
    for nombre in POSIBLES_EXCEL:
        print(f"- {nombre}")
    print("O ejecutá: python tools/agregarVictoriaBath.py ruta/del/excel.xlsx")
    raise SystemExit(1)


def leer_productos_excel(ruta_excel):
    wb = load_workbook(ruta_excel, data_only=True)
    hoja = wb.active

    encabezados = [limpiar_texto(celda.value).lower() for celda in hoja[1]]
    requeridas = ["nombre", "precio", "imagen", "categoria"]

    faltantes = [col for col in requeridas if col not in encabezados]
    if faltantes:
        print("El Excel debe tener estas columnas exactas en la primera fila:")
        print("nombre, precio, imagen, categoria")
        print(f"Faltan: {', '.join(faltantes)}")
        raise SystemExit(1)

    indices = {nombre: encabezados.index(nombre) for nombre in requeridas}
    productos_nuevos = []
    errores = []

    for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
        nombre = limpiar_texto(fila[indices["nombre"]])
        precio = fila[indices["precio"]]
        imagen = limpiar_texto(fila[indices["imagen"]])
        categoria = normalizar_categoria(fila[indices["categoria"]])

        if not any(fila):
            continue

        if not nombre or not imagen or not categoria:
            errores.append(numero_fila)
            continue

        productos_nuevos.append({
            "nombre": nombre,
            "precio": formatear_precio(precio),
            "imagen": imagen,
            "categoria": categoria,
            "tipoEspecial": "imagenGrande"
        })

    if errores:
        print("Estas filas se saltaron porque tienen nombre, imagen o categoría inválida:")
        print(", ".join(map(str, errores)))

    return productos_nuevos


def main():
    ruta_excel = buscar_excel()

    if not ARCHIVO_JSON.exists():
        print("No encontré productos.json en esta carpeta.")
        raise SystemExit(1)

    with ARCHIVO_JSON.open("r", encoding="utf-8") as archivo:
        productos = json.load(archivo)

    productos_nuevos = leer_productos_excel(ruta_excel)

    productos = [
        producto for producto in productos
        if producto.get("categoria") not in ["VICTORIA'S SECRET", "BATH AND BODY WORKS"]
    ]

    productos.extend(productos_nuevos)

    with ARCHIVO_JSON.open("w", encoding="utf-8") as archivo:
        json.dump(productos, archivo, ensure_ascii=False, indent=4)

    print(f"Productos agregados desde {ruta_excel}: {len(productos_nuevos)}")
    print("Categorías actualizadas: Victoria's Secret y Bath and Body Works.")


if __name__ == "__main__":
    main()
