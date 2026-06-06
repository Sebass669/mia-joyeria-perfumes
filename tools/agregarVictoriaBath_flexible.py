import json
import sys
import re
from pathlib import Path
from openpyxl import load_workbook

ARCHIVO_JSON = Path('productos.json')
CARPETA_IMAGENES = Path('imagers/victoria_bath')

POSIBLES_EXCEL = [
    'victoria_bath.xlsx',
    'victoria_secret_bath.xlsx',
    'productos_victoria_bath.xlsx',
    'productosEspeciales.xlsx',
    'productos_especiales.xlsx',
]

CATEGORIAS_VALIDAS = {
    "VICTORIA'S SECRET": "VICTORIA'S SECRET",
    'VICTORIA SECRET': "VICTORIA'S SECRET",
    'VICTORIAS SECRET': "VICTORIA'S SECRET",
    'VICTORIA´S SECRET': "VICTORIA'S SECRET",
    'VICTORIA’S SECRET': "VICTORIA'S SECRET",
    'BATH AND BODY WORKS': 'BATH AND BODY WORKS',
    'BATH & BODY WORKS': 'BATH AND BODY WORKS',
    'BATH BODY WORKS': 'BATH AND BODY WORKS',
}


def limpiar_texto(valor):
    return str(valor or '').strip()


def normalizar_encabezado(valor):
    return limpiar_texto(valor).lower().replace(' ', '')


def normalizar_categoria(valor):
    categoria = limpiar_texto(valor).upper()
    return CATEGORIAS_VALIDAS.get(categoria)


def formatear_precio(valor):
    if valor is None or str(valor).strip() == '':
        return 'Consultar precio'

    texto = str(valor).strip()
    if texto.lower() in ['consultar', 'consultar precio', 'sin precio']:
        return 'Consultar precio'

    texto_limpio = texto.replace('₡', '').replace('¢', '').replace(',', '').strip()
    try:
        numero = float(texto_limpio)
        return f'₡{numero:.2f}'
    except ValueError:
        return texto


def buscar_excel():
    if len(sys.argv) > 1:
        archivo = Path(sys.argv[1])
        if archivo.exists():
            return archivo
        raise SystemExit(f'No encontré el Excel indicado: {archivo}')

    for nombre in POSIBLES_EXCEL:
        archivo = Path(nombre)
        if archivo.exists():
            return archivo

    raise SystemExit('No encontré el Excel. Ponelo junto al script o pasá la ruta como argumento.')


def extension_imagen(img):
    formato = (getattr(img, 'format', '') or '').lower()
    if formato in ['jpeg', 'jpg']:
        return 'jpg'
    if formato == 'webp':
        return 'webp'
    if formato == 'png':
        return 'png'
    return 'png'


def buscar_fila_producto_para_imagen(fila_imagen, filas_productos, usadas):
    """Asocia la imagen a la fila de producto más cercana hacia abajo.
    Esto arregla Excel donde la imagen está anclada una o dos filas arriba del texto.
    """
    candidatas = [f for f in filas_productos if f not in usadas and f >= fila_imagen]
    if candidatas:
        fila = min(candidatas, key=lambda f: (f - fila_imagen, f))
        if fila - fila_imagen <= 5:
            return fila

    # Respaldo: buscar exacta/cercana hacia arriba por si la imagen quedó anclada debajo.
    candidatas = [f for f in filas_productos if f not in usadas]
    if candidatas:
        fila = min(candidatas, key=lambda f: (abs(f - fila_imagen), f))
        if abs(fila - fila_imagen) <= 5:
            return fila

    return None


def extraer_imagenes_por_fila(hoja, filas_productos):
    CARPETA_IMAGENES.mkdir(parents=True, exist_ok=True)
    imagenes = {}
    usadas = set()

    imagenes_ordenadas = sorted(
        hoja._images,
        key=lambda img: (img.anchor._from.row + 1, img.anchor._from.col + 1)
    )

    for i, img in enumerate(imagenes_ordenadas, start=1):
        fila_imagen = img.anchor._from.row + 1
        fila_producto = buscar_fila_producto_para_imagen(fila_imagen, filas_productos, usadas)

        if fila_producto is None:
            continue

        ext = extension_imagen(img)
        nombre_archivo = f'producto_fila_{fila_producto}_{i}.{ext}'
        ruta_salida = CARPETA_IMAGENES / nombre_archivo

        with open(ruta_salida, 'wb') as archivo:
            archivo.write(img._data())

        imagenes[fila_producto] = str(ruta_salida).replace('\\', '/')
        usadas.add(fila_producto)

    return imagenes


def leer_productos_excel(ruta_excel):
    wb = load_workbook(ruta_excel, data_only=True)
    hoja = wb.active

    encabezados = [normalizar_encabezado(celda.value) for celda in hoja[1]]
    requeridas = ['nombre', 'precio', 'categoria']

    faltantes = [col for col in requeridas if col not in encabezados]
    if faltantes:
        raise SystemExit(f'Faltan columnas: {", ".join(faltantes)}')

    indices = {nombre: encabezados.index(nombre) for nombre in requeridas}

    filas_productos = []
    datos_por_fila = {}
    categorias_invalidas = []

    for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
        if not any(fila):
            continue

        nombre = limpiar_texto(fila[indices['nombre']])
        precio = fila[indices['precio']]
        categoria_original = limpiar_texto(fila[indices['categoria']])
        categoria = normalizar_categoria(categoria_original)

        if not nombre:
            continue

        filas_productos.append(numero_fila)
        datos_por_fila[numero_fila] = {
            'nombre': nombre,
            'precio': precio,
            'categoria_original': categoria_original,
            'categoria': categoria,
        }

        if not categoria:
            categorias_invalidas.append((numero_fila, categoria_original))

    imagenes_por_fila = extraer_imagenes_por_fila(hoja, filas_productos)

    productos_nuevos = []
    filas_sin_imagen = []
    filas_sin_categoria_valida = []

    for numero_fila in filas_productos:
        datos = datos_por_fila[numero_fila]
        imagen = imagenes_por_fila.get(numero_fila, '')
        categoria = datos['categoria']

        if not imagen:
            filas_sin_imagen.append(numero_fila)
            continue

        if not categoria:
            filas_sin_categoria_valida.append((numero_fila, datos['categoria_original']))
            continue

        productos_nuevos.append({
            'nombre': datos['nombre'],
            'precio': formatear_precio(datos['precio']),
            'imagen': imagen,
            'categoria': categoria,
            'tipoEspecial': 'imagenGrande'
        })

    if filas_sin_imagen:
        print('Filas con producto pero sin imagen asociada:')
        print(', '.join(map(str, filas_sin_imagen)))

    if filas_sin_categoria_valida:
        print('Filas saltadas por categoría no válida:')
        for fila, categoria in filas_sin_categoria_valida:
            print(f'- Fila {fila}: {categoria}')
        print('Categorías válidas: VICTORIA\'S SECRET o BATH AND BODY WORKS')

    return productos_nuevos


def main():
    ruta_excel = buscar_excel()

    if not ARCHIVO_JSON.exists():
        raise SystemExit('No encontré productos.json en esta carpeta. Ejecutá el script desde la raíz del proyecto.')

    with ARCHIVO_JSON.open('r', encoding='utf-8') as archivo:
        productos = json.load(archivo)

    productos_nuevos = leer_productos_excel(ruta_excel)

    productos = [
        producto for producto in productos
        if producto.get('categoria') not in ["VICTORIA'S SECRET", 'BATH AND BODY WORKS']
    ]

    # Los pone de primeros en el catálogo.
    productos = productos_nuevos + productos

    with ARCHIVO_JSON.open('w', encoding='utf-8') as archivo:
        json.dump(productos, archivo, ensure_ascii=False, indent=4)

    print(f'Productos agregados desde {ruta_excel}: {len(productos_nuevos)}')
    print('Imágenes extraídas en:', CARPETA_IMAGENES)


if __name__ == '__main__':
    main()
